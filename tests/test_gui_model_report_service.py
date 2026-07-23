"""Проверки структуры Excel-отчёта модели."""

from types import SimpleNamespace

from openpyxl import load_workbook

from gui.app_state import GraphNode, Model, NodeKind, NodeStatus, Variant
from gui.services import model_report_service as report_svc


def test_excel_report_has_model_and_selected_experiment_sheets(tmp_path):
    composition = SimpleNamespace(
        composition={"C1": 0.8, "C2": 0.2},
        composition_data={
            "MW": {"C1": 16.04, "C2": 30.07},
            "Tc": {"C1": 190.6, "C2": 305.3},
        },
    )
    variant = Variant(variant_id="base", title="Base", composition=composition)
    variant.nodes["exp_1"] = GraphNode(
        node_id="exp_1", kind=NodeKind.EXPERIMENT, title="DLE",
        status=NodeStatus.FRESH,
        params={"kind": "dle", "T_c": 80.0, "pressures": [300.0, 200.0]},
        result={"columns": ["pressure", "Bo"], "rows": [[300.0, 1.2]]},
    )
    model = Model(model_id="fluid_a", title="Fluid A", project_id="project_a",
                  field_name="Field", eos="BRSEOS", n_components=2,
                  loaded=True, variants={"base": variant})

    path = report_svc.export_model_report(
        tmp_path / "report", model, variant, str(tmp_path / "models.json"),
        {"composition": True, "dle": True, "cce": False, "separator": False,
         "flash": False, "envelope": False, "lab_data": False},
    )

    book = load_workbook(path, data_only=True)
    assert book.sheetnames == ["Model", "DLE 1"]
    assert book["Model"][1][0].value == "Property"
    assert book["DLE 1"][1][0].value == "Parameter"
    assert book["DLE 1"][8][0].value == "Pressure, bar"
    assert book["DLE 1"][9][1].value == 1.2
