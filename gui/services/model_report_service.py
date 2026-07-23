"""Excel report for one model and its completed calculations."""

from __future__ import annotations

from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

from gui.app_state import Model, NodeKind, Variant
from gui.services import flash_service
from gui.services import lab_data_service as lab_svc

_HEADER_FILL = PatternFill("solid", fgColor="C2CFDB")
_PROPERTY_LABELS = {
    "molecular_ weight": "Molar mass, g/mol",
    "molar_volume": "Molar volume",
    "molar_density": "Molar density",
    "density": "Density",
    "z": "Z-factor",
    "viscosity": "Viscosity",
}
_COLUMN_LABELS = {
    "pressure": "Pressure, bar",
    "temperature": "Temperature, C",
    "Bo": "Oil volume factor, Bo",
    "Rs": "Solution gas-oil ratio, Rs",
    "v/v_sat": "Relative volume, V/Vsat",
    "v/v_res": "Relative volume, V/Vres",
    "Compressibility": "Compressibility, 1/bar",
    "liquid_density": "Liquid density",
    "vapor_density": "Vapor density",
    "liquid_z": "Liquid Z-factor",
    "vapor_z": "Vapor Z-factor",
    "vapor_mole_frac": "Vapor mole fraction",
    "liquid_mole_frac": "Liquid mole fraction",
    "MW": "Molar mass, g/mol",
    "Tc": "Critical temperature, K",
    "Pc": "Critical pressure, bar",
}


def _header(value: object) -> str:
    return _COLUMN_LABELS.get(str(value), str(value))


def _safe_sheet_name(base: str, used: set[str]) -> str:
    cleaned = "".join("_" if char in r"[]:*?/\\" else char for char in base).strip()
    cleaned = cleaned[:31] or "Report"
    candidate = cleaned
    index = 2
    while candidate in used:
        suffix = f" {index}"
        candidate = cleaned[:31 - len(suffix)] + suffix
        index += 1
    used.add(candidate)
    return candidate


def _write_table(sheet, start_row: int, columns: list[str], rows: list[list[Any]]) -> int:
    for column, value in enumerate(columns, start=1):
        cell = sheet.cell(start_row, column, value)
        cell.font = Font(bold=True)
        cell.fill = _HEADER_FILL
    for row_index, row in enumerate(rows, start=start_row + 1):
        for column, value in enumerate(row, start=1):
            sheet.cell(row_index, column, value)
    sheet.freeze_panes = f"A{start_row + 1}"
    sheet.auto_filter.ref = (
        f"A{start_row}:{chr(64 + min(len(columns), 26))}{start_row + len(rows)}"
    )
    for column in sheet.columns:
        letter = column[0].column_letter
        width = min(38, max(12, max(len(str(cell.value or "")) for cell in column) + 2))
        sheet.column_dimensions[letter].width = width
    return start_row + len(rows) + 2


def _model_sheet(book: Workbook, model: Model, variant: Variant,
                 include_composition: bool) -> None:
    sheet = book.active
    sheet.title = "Model"
    comp = variant.composition
    metadata: list[list[Any]] = [
        ["Model", model.title],
        ["Model id", model.model_id],
        ["Field", model.field_name or ""],
        ["EOS", model.eos or ""],
        ["Components", model.n_components],
    ]
    _write_table(sheet, 1, ["Property", "Value"], metadata)
    if comp is None or not include_composition:
        return
    data = comp.composition_data if isinstance(comp.composition_data, dict) else {}
    property_keys = sorted({key for value in data.values() if isinstance(value, dict)
                            for key in value if key not in {"name", "Name"}})
    rows = []
    for name, fraction in comp.composition.items():
        props = data.get(name, {}) if isinstance(data.get(name), dict) else {}
        rows.append([name, float(fraction)] + [props.get(key) for key in property_keys])
    _write_table(sheet, len(metadata) + 4,
                 ["Component", "Mole fraction"] + [_header(key) for key in property_keys], rows)


def _flash_rows(result) -> tuple[list[str], list[list[Any]], list[list[Any]]]:
    main = [["Phase mole fraction", result.vapor.mole_fraction, result.liquid.mole_fraction]]
    main.extend([[label, result.vapor.properties.get(key), result.liquid.properties.get(key)]
                 for key, label in flash_service.PHASE_PROPERTY_ROWS])
    names = list((result.liquid.composition or result.vapor.composition or {}).keys())
    composition = []
    for name in names:
        vapor = result.vapor.composition.get(name) if result.vapor.composition else None
        liquid = result.liquid.composition.get(name) if result.liquid.composition else None
        k_value = vapor / liquid if vapor is not None and liquid not in (None, 0) else None
        composition.append([name, vapor, liquid, k_value])
    return ["Property", "Vapor", "Liquid"], main, composition


def _write_node_sheet(book: Workbook, used: set[str], node, model: Model,
                      db_path: str, include_lab_data: bool) -> None:
    sequence = node.node_id.split("_")[-1]
    if node.kind is NodeKind.EXPERIMENT:
        base = f"{str(node.params.get('kind', 'Experiment')).upper()} {sequence}"
    elif node.kind is NodeKind.PHASE_ENVELOPE:
        base = f"Envelope {sequence}"
    else:
        base = f"Flash {sequence}"
    sheet = book.create_sheet(_safe_sheet_name(str(node.params.get("label") or base), used))
    details = [["Calculation", node.title], ["Status", node.status.name]]
    details.extend([[str(key), value] for key, value in node.params.items()
                    if key not in {"lab_data", "lab_data_ref"}])
    next_row = _write_table(sheet, 1, ["Parameter", "Value"], details)

    if node.kind is NodeKind.FLASH:
        columns, rows, composition = _flash_rows(node.result)
        next_row = _write_table(sheet, next_row, columns, rows)
        if composition:
            _write_table(sheet, next_row,
                         ["Component", "Vapor yi", "Liquid xi", "K = yi/xi"], composition)
        return
    if node.kind is NodeKind.EXPERIMENT:
        result = node.result if isinstance(node.result, dict) else {}
        _write_table(sheet, next_row, [_header(column) for column in result.get("columns", [])],
                     [list(row) for row in result.get("rows", [])])
        if include_lab_data:
            raw = lab_svc.effective_lab_data(db_path, model.project_id, model.model_id,
                                             node.params)
            if isinstance(raw, dict) and raw.get("columns"):
                next_row += len(result.get("rows", [])) + 3
                _write_table(sheet, next_row, [_header(column) for column in raw["columns"]],
                             [list(row) for row in raw.get("rows", [])])
        return
    result = node.result if isinstance(node.result, dict) else {}
    table = result.get("table", {})
    _write_table(sheet, next_row, [_header(column) for column in table.get("columns", [])],
                 [list(row) for row in table.get("rows", [])])


def export_model_report(path: str | Path, model: Model, variant: Variant,
                        db_path: str, options: dict[str, bool]) -> str:
    """Writes a workbook from the current model snapshot and returns its path."""
    destination = Path(path).with_suffix(".xlsx")
    book = Workbook()
    _model_sheet(book, model, variant, options.get("composition", True))
    used = {"Model"}
    enabled_experiments = {kind for kind in ("dle", "cce", "separator")
                           if options.get(kind, True)}
    for node in variant.nodes.values():
        if node.result is None:
            continue
        if node.kind is NodeKind.FLASH and not options.get("flash", True):
            continue
        if node.kind is NodeKind.PHASE_ENVELOPE and not options.get("envelope", True):
            continue
        if (node.kind is NodeKind.EXPERIMENT
                and str(node.params.get("kind")) not in enabled_experiments):
            continue
        if node.kind not in {NodeKind.FLASH, NodeKind.EXPERIMENT, NodeKind.PHASE_ENVELOPE}:
            continue
        _write_node_sheet(book, used, node, model, db_path,
                          options.get("lab_data", True))
    book.save(destination)
    return str(destination)
